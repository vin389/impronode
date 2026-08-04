import json
import ast
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception:  # pragma: no cover - handled at runtime via clear error
    Workbook = None
    load_workbook = None

SCHEMA_VERSION = "1.0"


class ProjectFormatError(Exception):
    pass


def _require_openpyxl() -> None:
    if Workbook is None or load_workbook is None:
        raise RuntimeError(
            "openpyxl is required for XLSX project files. Install with: pip install openpyxl"
        )


def _coerce_scalar(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (bool, int, float)):
        return value

    text = str(value).strip()
    if text == "":
        return None

    lower = text.lower()
    if lower == "true":
        return True
    if lower == "false":
        return False

    try:
        if text.isdigit() or (text.startswith("-") and text[1:].isdigit()):
            return int(text)
        return float(text)
    except ValueError:
        pass

    if text[:1] in "[{\"":
        try:
            return json.loads(text)
        except Exception:
            return text
    return text


def _compact_text(value: Any, limit: int = 120) -> str:
    text = str(value).replace("\r", " ").replace("\n", " ").strip()
    if len(text) <= limit:
        return text
    return text[: max(0, limit - 3)] + "..."


def _parse_params_object(raw_params: str) -> tuple[dict[str, Any] | None, str | None]:
    """
    Parse params_json into a dict.

    Returns:
      (dict, None) on success
      (None, error_message) on failure
    """
    text = str(raw_params).strip()
    if not text:
        return {}, None

    # 1) Canonical JSON object format.
    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed, None
        if isinstance(parsed, str):
            # Some legacy files may contain a JSON-encoded JSON string.
            nested = json.loads(parsed)
            if isinstance(nested, dict):
                return nested, None
        return None, f"params_json must be a JSON object/dict; got {type(parsed).__name__}"
    except Exception as e_json:
        json_error = str(e_json)

    # 2) Common Excel/manual-edit case: wrapping single quotes around the whole JSON text.
    if len(text) >= 2 and text[0] == "'" and text[-1] == "'":
        inner = text[1:-1].strip()
        try:
            parsed = json.loads(inner)
            if isinstance(parsed, dict):
                return parsed, None
        except Exception:
            pass

    # 2.5) Recover common truncation case where file_list tail is cut and
    #      breaks JSON (for example: Unterminated string ... near file_list).
    if text.startswith("{") and '"file_list"' in text:
        pivot = text.find(',"file_list"')
        if pivot == -1:
            pivot = text.find('"file_list"')
        if pivot > 0:
            trimmed = text[:pivot].rstrip().rstrip(",") + "}"
            try:
                parsed = json.loads(trimmed)
                if isinstance(parsed, dict):
                    return parsed, None
            except Exception:
                pass

    # 3) Legacy Python-literal dict representation.
    try:
        lit = ast.literal_eval(text)
        if isinstance(lit, dict):
            return lit, None
    except Exception:
        pass

    # 4) Legacy malformed Windows-path blobs where backslashes were not escaped.
    #    Example: {"pattern":"D:\images\..."} saved without JSON escaping.
    escaped_text = text.replace("\\", "\\\\")
    if escaped_text != text:
        try:
            parsed = json.loads(escaped_text)
            if isinstance(parsed, dict):
                return parsed, None
        except Exception:
            pass
        try:
            lit = ast.literal_eval(escaped_text)
            if isinstance(lit, dict):
                return lit, None
        except Exception:
            pass

    return None, f"invalid params_json ({json_error})"


def save_project(
    file_path: str | Path,
    nodes: list[dict[str, Any]],
    links: list[dict[str, str]],
    app_meta: dict[str, Any] | None = None,
) -> None:
    _require_openpyxl()

    wb = Workbook()

    ws_meta = wb.active
    ws_meta.title = "ProjectMeta"
    ws_meta.append(["key", "value"])
    ws_meta.append(["schema_version", SCHEMA_VERSION])
    if app_meta:
        for key, value in app_meta.items():
            if key == "extra_tables":
                continue
            ws_meta.append([str(key), json.dumps(value, ensure_ascii=False)])

    ws_nodes = wb.create_sheet("Nodes")
    fixed_cols = [
        "node_id",
        "node_type",
        "node_name",
        "x",
        "y",
        "width",
        "height",
        "params_json",
    ]

    # Keep Nodes schema compact: params_json is the single canonical params field.
    ws_nodes.append(fixed_cols)

    for n in nodes:
        row = {
            "node_id": n.get("node_id", ""),
            "node_type": n.get("node_type", ""),
            "node_name": n.get("node_name", ""),
            "x": n.get("x", 0),
            "y": n.get("y", 0),
            "width": n.get("width", 0),
            "height": n.get("height", 0),
            "params_json": json.dumps(n.get("params", {}), ensure_ascii=False),
        }

        ws_nodes.append([row.get(c, "") for c in fixed_cols])

    ws_links = wb.create_sheet("Links")
    ws_links.append(["src_node", "src_pin", "dst_node", "dst_pin"])
    for lk in links:
        ws_links.append([
            lk.get("src_node", ""),
            lk.get("src_pin", ""),
            lk.get("dst_node", ""),
            lk.get("dst_pin", ""),
        ])

    # Optional future-facing tables for large array-like data exports/imports.
    extra_tables = (app_meta or {}).get("extra_tables", {}) if app_meta else {}
    if isinstance(extra_tables, dict):
        for sheet_name, rows in extra_tables.items():
            sname = str(sheet_name)
            if not sname or sname in ("ProjectMeta", "Nodes", "Links"):
                continue
            if sname in wb.sheetnames:
                continue
            ws = wb.create_sheet(sname)
            if not isinstance(rows, list) or not rows:
                continue
            if not isinstance(rows[0], dict):
                continue

            cols = sorted({str(k) for r in rows if isinstance(r, dict) for k in r.keys()})
            ws.append(cols)
            for r in rows:
                if not isinstance(r, dict):
                    continue
                ws.append([r.get(c, "") for c in cols])

    wb.save(str(file_path))


def load_project(file_path: str | Path) -> tuple[list[dict[str, Any]], list[dict[str, str]], dict[str, Any]]:
    _require_openpyxl()

    wb = load_workbook(str(file_path), data_only=True)
    if "Nodes" not in wb.sheetnames:
        raise ProjectFormatError("Missing required worksheet: Nodes")
    if "Links" not in wb.sheetnames:
        raise ProjectFormatError("Missing required worksheet: Links")

    meta: dict[str, Any] = {}
    issues: list[str] = []
    if "ProjectMeta" in wb.sheetnames:
        ws_meta = wb["ProjectMeta"]
        for row in ws_meta.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            key = str(row[0])
            val = row[1]
            try:
                meta[key] = json.loads(val) if isinstance(val, str) else val
            except Exception:
                meta[key] = val

    ws_nodes = wb["Nodes"]
    header = [str(c) if c is not None else "" for c in next(ws_nodes.iter_rows(min_row=1, max_row=1, values_only=True))]
    col_idx = {name: idx for idx, name in enumerate(header)}

    required = ["node_id", "node_type", "x", "y", "width", "height", "params_json"]
    for req in required:
        if req not in col_idx:
            raise ProjectFormatError(f"Nodes sheet missing required column: {req}")

    dynamic_param_cols = [h for h in header if h.startswith("param_")]

    nodes: list[dict[str, Any]] = []
    seen_node_ids: set[str] = set()
    for row_no, row in enumerate(ws_nodes.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue

        node_id = row[col_idx["node_id"]]
        node_type = row[col_idx["node_type"]]
        if node_id is None or node_type is None:
            issues.append(f"Nodes row {row_no}: missing node_id or node_type; row skipped.")
            continue

        node_id_str = str(node_id)
        if node_id_str in seen_node_ids:
            issues.append(f"Nodes row {row_no}: duplicate node_id '{node_id_str}'.")
        seen_node_ids.add(node_id_str)

        params: dict[str, Any] = {}
        raw_params = row[col_idx["params_json"]]
        if isinstance(raw_params, str) and raw_params.strip():
            parsed_params, parse_error = _parse_params_object(raw_params)
            if parsed_params is not None:
                params.update(parsed_params)
            else:
                issues.append(
                    "Nodes row "
                    f"{row_no} (node_id='{node_id_str}', node_type='{node_type}'): "
                    "invalid params_json; ignored. "
                    f"error='{_compact_text(parse_error or 'unknown')}' "
                    f"value='{_compact_text(raw_params)}'."
                )
        elif raw_params not in (None, ""):
            issues.append(
                "Nodes row "
                f"{row_no} (node_id='{node_id_str}', node_type='{node_type}'): "
                "params_json is not text; ignored. "
                f"type={type(raw_params).__name__} value='{_compact_text(raw_params)}'."
            )

        for pname in dynamic_param_cols:
            idx = col_idx[pname]
            if idx >= len(row):
                continue
            v = row[idx]
            if v is None:
                continue
            key = pname[len("param_"):]
            coerced = _coerce_scalar(v)
            if coerced is not None:
                params[key] = coerced

        try:
            x = int(float(row[col_idx["x"]] or 0))
            y = int(float(row[col_idx["y"]] or 0))
            width = int(float(row[col_idx["width"]] or 0))
            height = int(float(row[col_idx["height"]] or 0))
        except Exception:
            issues.append(f"Nodes row {row_no}: invalid numeric geometry; row skipped.")
            continue

        if width <= 0 or height <= 0:
            issues.append(f"Nodes row {row_no}: non-positive width/height; defaults may apply.")

        nodes.append({
            "node_id": node_id_str,
            "node_type": str(node_type),
            "node_name": str(row[col_idx["node_name"]]) if "node_name" in col_idx and row[col_idx["node_name"]] is not None else "",
            "x": x,
            "y": y,
            "width": width,
            "height": height,
            "params": params,
        })

    ws_links = wb["Links"]
    links: list[dict[str, str]] = []
    for row_no, row in enumerate(ws_links.iter_rows(min_row=2, values_only=True), start=2):
        if not row:
            continue
        src_node, src_pin, dst_node, dst_pin = (row + (None, None, None, None))[:4]
        if None in (src_node, src_pin, dst_node, dst_pin):
            issues.append(f"Links row {row_no}: missing required column value; row skipped.")
            continue
        src_node_str = str(src_node)
        dst_node_str = str(dst_node)
        if src_node_str not in seen_node_ids or dst_node_str not in seen_node_ids:
            issues.append(
                f"Links row {row_no}: references unknown node(s) '{src_node_str}' or '{dst_node_str}'."
            )
        links.append({
            "src_node": src_node_str,
            "src_pin": str(src_pin),
            "dst_node": dst_node_str,
            "dst_pin": str(dst_pin),
        })

    extra_tables: dict[str, list[dict[str, Any]]] = {}
    for sname in wb.sheetnames:
        if sname in ("ProjectMeta", "Nodes", "Links"):
            continue
        ws = wb[sname]
        rows_iter = ws.iter_rows(min_row=1, max_row=1, values_only=True)
        first = next(rows_iter, None)
        if not first:
            extra_tables[sname] = []
            continue
        headers = [str(h) if h is not None else "" for h in first]
        table_rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            row_dict: dict[str, Any] = {}
            for idx, h in enumerate(headers):
                if not h:
                    continue
                v = row[idx] if idx < len(row) else None
                if v is not None:
                    row_dict[h] = v
            if row_dict:
                table_rows.append(row_dict)
        extra_tables[sname] = table_rows

    if issues:
        meta["load_issues"] = issues
    if extra_tables:
        meta["extra_tables"] = extra_tables

    return nodes, links, meta
